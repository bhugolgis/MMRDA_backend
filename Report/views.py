from rest_framework.generics import ListAPIView
from .serializers import *
from SocialMonitoring.models import *
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework import generics
from rest_framework.views import APIView
from Training.models import *
from openpyxl import Workbook
from rest_framework.viewsets import ReadOnlyModelViewSet
from drf_excel.mixins import XLSXFileMixin
from drf_excel.renderers import XLSXRenderer
from django_filters.rest_framework import DjangoFilterBackend
import pandas as pd
from django.http import StreamingHttpResponse
from django.http import HttpResponse
from django.http import JsonResponse
import calendar
# from django_excel_response import ExcelResponse



''' --------------------------Labour Camp Report View----------------------------'''


# The `LabourcampReportPackageView` class is a view in a Django REST framework that retrieves previous
# and latest data for a given package and labour camp name.
class LabourcampReportPackageView(ListAPIView):
    serializer_class = LabourcampReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, labourCampName ,*args, **kwargs):
        """
        This function retrieves the previous and latest data for a given package and labour camp name.

        """
        try:
            # previous = LabourCamp.objects.filter(packages=packages, labourCampName=labourCampName ).order_by('-id')[1:]
            previous = LabourCamp.objects.filter(packages=packages, labourCampName=labourCampName )

            if not previous:
                return JsonResponse({'status':'error','message':'Data Not Found previous'}, status=400)
           
            latest = LabourCamp.objects.filter(packages=packages, labourCampName=labourCampName).latest('id')

            if not latest:
                return JsonResponse({'status':'error','message':'Data Not Found'}, status=400 )           
            # latest_serializer = LabourcampReportSerializer(latest).data
           
            previous_data = self.serializer_class(previous, many=True).data
            print(type(previous_data))
            latest_data = self.serializer_class(latest).data

            latest_data['properties']['id'] = latest_data['id'] 
            for feature in previous_data['features']:
                feature['properties']['id'] = feature['id']           

            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            'Previous': previous_data,
                             'latest': latest_data}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'Message': 'There is no data available for this Package or Quarter',
                            'status' : 'Failed'}, status=400)


# The `LabourCampReportQuarterView` class is a Django view that retrieves data from the `LabourCamp`
# model based on the specified quarter, labour camp name, and year, and returns the previous and
# latest data in a response.LabourCampReportQuarterView
class LabourCampReportQuarterView(generics.ListAPIView):

    serializer_class = LabourcampReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, quarter, labourCampName, year, *args, **kwarges):
        """
        This function retrieves data from the LabourCamp model based on the specified quarter, labour
        camp name, and year, and returns the previous and latest data in a response.
        
        """
        try:
            previous = LabourCamp.objects.filter(quarter=quarter, labourCampName=labourCampName, dateOfMonitoring__year=year).order_by('-id')[1:]

            latest = LabourCamp.objects.filter( quarter=quarter, labourCampName=labourCampName , dateOfMonitoring__year=year).latest('id')
            previousData = self.serializer_class(previous, many=True).data
            latestData = self.serializer_class(latest).data


            latestData['properties']['id'] = latestData['id'] 
            for feature in previousData['features']:
                feature['properties']['id'] = feature['id']  

            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            'Previous': previousData,
                            'latest': latestData},
                            status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available',
                            'status' : 'Failed'}, status=400)
            

class labourcampreportpackageExcelDownloadView(generics.ListAPIView):
    serializer_class = LabourcampExcelReportSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages', 'labourCampName']

    def get_queryset(self):
        queryset = LabourCamp.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values(
            'id','quarter', 'packages','dateOfMonitoring', 'labourCampName', 'labourCampId',
                  'isToilet', 'toiletCondition','toiletPhotograph','toiletRemarks',
                  'isDrinkingWater','drinkingWaterCondition' ,'drinkingWaterPhotographs','drinkingWaterRemarks',
                    'isDemarkationOfPathways','demarkationOfPathwaysCondition','demarkationOfPathwaysPhotographs','demarkationOfPathwaysRemark' ,
                    'isSignagesLabeling','signagesLabelingCondition' ,'signagesLabelingPhotographs','signagesLabelingRemarks',
                    'isKitchenArea','kitchenAreaCondition','kitchenAreaPhotographs','kitchenAreaRemarks',
                    'isFireExtinguish','fireExtinguishCondition','fireExtinguishPhotographs','fireExtinguishRemarks',
                     'isRoomsOrDoms' ,'roomsOrDomsCondition','roomsOrDomsPhotographs' ,'roomsOrDomsRemarks',
                     'isSegregationOfWaste','segregationOfWasteCondition','segregationOfWastePhotographs','segregationOfWasteRemarks',
                    'isRegularHealthCheckup','regularHealthCheckupCondition','regularHealthCheckupPhotographs','regularHealthCheckupRemarks',
                     'isAvailabilityOfDoctor', 'availabilityOfDoctorCondition','availabilityOfDoctorPhotographs','availabilityOfDoctorRemarks',
                      'isFirstAidKit','firstAidKitCondition' ,'firstAidKitPhotographs','firstAidKitRemarks',
                    'transportationFacility', 'modeOfTransportation',
                    'photographs' ,'documents','remarks'
            # Add more fields as needed
        )
        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfMonitoring': 'Monitoring Date',
                'labourCampName': 'Labour Camp Name',
                'labourCampId': 'Labour Camp ID',
                'isToilet': 'Is Toilet Available',
                'toiletCondition': 'Toilet Condition',
                'toiletPhotograph': 'Toilet Photographs',
                'toiletRemarks': 'Toilet Remarks',
                'isDrinkingWater': 'Is Drinking Water Available',
                'drinkingWaterCondition': 'Drinking Water Condition',
                'drinkingWaterPhotographs': 'Drinking Water Photographs',
                'drinkingWaterRemarks': 'Drinking Water Remarks',
                'isDemarkationOfPathways': 'Is Pathways Demarked',
                'demarkationOfPathwaysCondition': 'Pathways Demarkation Condition',
                'demarkationOfPathwaysPhotographs': 'Pathways Demarkation Photographs',
                'demarkationOfPathwaysRemark': 'Pathways Demarkation Remarks',
                'isSignagesLabeling': 'Is Signages and Labeling Available',
                'signagesLabelingCondition': 'Signages and Labeling Condition',
                'signagesLabelingPhotographs': 'Signages and Labeling Photographs',
                'signagesLabelingRemarks': 'Signages and Labeling Remarks',
                'isKitchenArea': 'Is Kitchen Area Available',
                'kitchenAreaCondition': 'Kitchen Area Condition',
                'kitchenAreaPhotographs': 'Kitchen Area Photographs',
                'kitchenAreaRemarks': 'Kitchen Area Remarks',
                'isFireExtinguish': 'Is Fire Extinguisher Available',
                'fireExtinguishCondition': 'Fire Extinguisher Condition',
                'fireExtinguishPhotographs': 'Fire Extinguisher Photographs',
                'fireExtinguishRemarks': 'Fire Extinguisher Remarks',
                'isRoomsOrDoms': 'Is Rooms/Dorms Available',
                'roomsOrDomsCondition': 'Rooms/Dorms Condition',
                'roomsOrDomsPhotographs': 'Rooms/Dorms Photographs',
                'roomsOrDomsRemarks': 'Rooms/Dorms Remarks',
                'isSegregationOfWaste': 'Is Waste Segregation Available',
                'segregationOfWasteCondition': 'Waste Segregation Condition',
                'segregationOfWastePhotographs': 'Waste Segregation Photographs',
                'segregationOfWasteRemarks': 'Waste Segregation Remarks',
                'isRegularHealthCheckup': 'Is Regular Health Checkup Done',
                'regularHealthCheckupCondition': 'Health Checkup Condition',
                'regularHealthCheckupPhotographs': 'Health Checkup Photographs',
                'regularHealthCheckupRemarks': 'Health Checkup Remarks',
                'isAvailabilityOfDoctor': 'Is Doctor Available',
                'availabilityOfDoctorCondition': 'Doctor Availability Condition',
                'availabilityOfDoctorPhotographs': 'Doctor Availability Photographs',
                'availabilityOfDoctorRemarks': 'Doctor Availability Remarks',
                'isFirstAidKit': 'Is First Aid Kit Available',
                'firstAidKitCondition': 'First Aid Kit Condition',
                'firstAidKitPhotographs': 'First Aid Kit Photographs',
                'firstAidKitRemarks': 'First Aid Kit Remarks',
                'transportationFacility': 'Transportation Facility',
                'modeOfTransportation': 'Mode of Transportation',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=labour_camp_report.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


import django_filters

class LabourCampFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')

    class Meta:
        model = LabourCamp
        fields = ['quarter', 'year', 'labourCampName']



# class labourQuarterExcelDownload(generics.ListAPIView):
class labourQuarterExcelDownload(generics.ListAPIView):
    serializer_class = LabourcampExcelReportSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'labourCampName']
    filterset_class = LabourCampFilter

    def get_queryset(self):
        queryset = LabourCamp.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values(
            'id','quarter', 'packages','dateOfMonitoring', 'labourCampName', 'labourCampId',
                  'isToilet', 'toiletCondition','toiletPhotograph','toiletRemarks',
                  'isDrinkingWater','drinkingWaterCondition' ,'drinkingWaterPhotographs','drinkingWaterRemarks',
                    'isDemarkationOfPathways','demarkationOfPathwaysCondition','demarkationOfPathwaysPhotographs','demarkationOfPathwaysRemark' ,
                    'isSignagesLabeling','signagesLabelingCondition' ,'signagesLabelingPhotographs','signagesLabelingRemarks',
                    'isKitchenArea','kitchenAreaCondition','kitchenAreaPhotographs','kitchenAreaRemarks',
                    'isFireExtinguish','fireExtinguishCondition','fireExtinguishPhotographs','fireExtinguishRemarks',
                     'isRoomsOrDoms' ,'roomsOrDomsCondition','roomsOrDomsPhotographs' ,'roomsOrDomsRemarks',
                     'isSegregationOfWaste','segregationOfWasteCondition','segregationOfWastePhotographs','segregationOfWasteRemarks',
                    'isRegularHealthCheckup','regularHealthCheckupCondition','regularHealthCheckupPhotographs','regularHealthCheckupRemarks',
                     'isAvailabilityOfDoctor', 'availabilityOfDoctorCondition','availabilityOfDoctorPhotographs','availabilityOfDoctorRemarks',
                      'isFirstAidKit','firstAidKitCondition' ,'firstAidKitPhotographs','firstAidKitRemarks',
                    'transportationFacility', 'modeOfTransportation',
                    'photographs' ,'documents','remarks'
            # Add more fields as needed
        )
        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfMonitoring': 'Monitoring Date',
                'labourCampName': 'Labour Camp Name',
                'labourCampId': 'Labour Camp ID',
                'isToilet': 'Is Toilet Available',
                'toiletCondition': 'Toilet Condition',
                'toiletPhotograph': 'Toilet Photographs',
                'toiletRemarks': 'Toilet Remarks',
                'isDrinkingWater': 'Is Drinking Water Available',
                'drinkingWaterCondition': 'Drinking Water Condition',
                'drinkingWaterPhotographs': 'Drinking Water Photographs',
                'drinkingWaterRemarks': 'Drinking Water Remarks',
                'isDemarkationOfPathways': 'Is Pathways Demarked',
                'demarkationOfPathwaysCondition': 'Pathways Demarkation Condition',
                'demarkationOfPathwaysPhotographs': 'Pathways Demarkation Photographs',
                'demarkationOfPathwaysRemark': 'Pathways Demarkation Remarks',
                'isSignagesLabeling': 'Is Signages and Labeling Available',
                'signagesLabelingCondition': 'Signages and Labeling Condition',
                'signagesLabelingPhotographs': 'Signages and Labeling Photographs',
                'signagesLabelingRemarks': 'Signages and Labeling Remarks',
                'isKitchenArea': 'Is Kitchen Area Available',
                'kitchenAreaCondition': 'Kitchen Area Condition',
                'kitchenAreaPhotographs': 'Kitchen Area Photographs',
                'kitchenAreaRemarks': 'Kitchen Area Remarks',
                'isFireExtinguish': 'Is Fire Extinguisher Available',
                'fireExtinguishCondition': 'Fire Extinguisher Condition',
                'fireExtinguishPhotographs': 'Fire Extinguisher Photographs',
                'fireExtinguishRemarks': 'Fire Extinguisher Remarks',
                'isRoomsOrDoms': 'Is Rooms/Dorms Available',
                'roomsOrDomsCondition': 'Rooms/Dorms Condition',
                'roomsOrDomsPhotographs': 'Rooms/Dorms Photographs',
                'roomsOrDomsRemarks': 'Rooms/Dorms Remarks',
                'isSegregationOfWaste': 'Is Waste Segregation Available',
                'segregationOfWasteCondition': 'Waste Segregation Condition',
                'segregationOfWastePhotographs': 'Waste Segregation Photographs',
                'segregationOfWasteRemarks': 'Waste Segregation Remarks',
                'isRegularHealthCheckup': 'Is Regular Health Checkup Done',
                'regularHealthCheckupCondition': 'Health Checkup Condition',
                'regularHealthCheckupPhotographs': 'Health Checkup Photographs',
                'regularHealthCheckupRemarks': 'Health Checkup Remarks',
                'isAvailabilityOfDoctor': 'Is Doctor Available',
                'availabilityOfDoctorCondition': 'Doctor Availability Condition',
                'availabilityOfDoctorPhotographs': 'Doctor Availability Photographs',
                'availabilityOfDoctorRemarks': 'Doctor Availability Remarks',
                'isFirstAidKit': 'Is First Aid Kit Available',
                'firstAidKitCondition': 'First Aid Kit Condition',
                'firstAidKitPhotographs': 'First Aid Kit Photographs',
                'firstAidKitRemarks': 'First Aid Kit Remarks',
                'transportationFacility': 'Transportation Facility',
                'modeOfTransportation': 'Mode of Transportation',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=labour_camp_report.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response
    

# ----------------------------------------------------------------

class ConstructionCampReportPackageView(ListAPIView):
    serializer_class = ConstructionCampReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, constructionSiteName,*args, **kwargs):
        try:
            previous = ConstructionSiteDetails.objects.filter(packages=packages, constructionSiteName=constructionSiteName  ).order_by('-id')[1:]
            latest = ConstructionSiteDetails.objects.filter(packages=packages, constructionSiteName=constructionSiteName).latest('id')
            previous_data = ConstructionCampReportSerializer(previous, many=True).data
            latest_data = ConstructionCampReportSerializer(latest).data

            latest_data['properties']['id'] = latest_data['id'] 
            for feature in previous_data['features']:
                feature['properties']['id'] = feature['id'] 

            

            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            'Previous': previous_data,
                            'latest': latest_data},
                            status=status.HTTP_200_OK)

        except Exception:
            return Response({'Message': 'There is no data available for the Package or Quarter',
                            'status' : 'Failed'}, status=400)


class ConstructionCampReportQuarterView(ListAPIView):
    serializer_class = ConstructionCampReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, quarter, constructionSiteName, year ,*args, **kwargs):
        try:
            previous = ConstructionSiteDetails.objects.filter(quarter=quarter, constructionSiteName=constructionSiteName ,dateOfMonitoring__year=year).order_by('-id')[1:]
            latest = latest = ConstructionSiteDetails.objects.filter(quarter=quarter, constructionSiteName=constructionSiteName , dateOfMonitoring__year=year).latest('id')
            previousData = self.serializer_class(previous, many=True).data
            latestData = self.serializer_class(latest).data

            latestData['properties']['id'] = latestData['id'] 
            for feature in previousData['features']:
                feature['properties']['id'] = feature['id'] 

            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            'Previous': previousData,
                            'latest': latestData},
                            status=status.HTTP_200_OK)
        except Exception:
            return Response({'Message': 'There is no data available for the Package or Quarter',
                            'status' : 'Failed'}, status=status.HTTP_400_BAD_REQUEST)


class ConstructionCampReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = ConstructionCampReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages', 'constructionSiteName']
    # filterset_class = LabourCampFilter

    def get_queryset(self):
        queryset = ConstructionSiteDetails.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values(
                    'id','quarter', 'packages','dateOfMonitoring' ,'constructionSiteName' , 'constructionSiteId',
                    'isDemarkationOfPathways','demarkationOfPathwaysCondition','demarkationOfPathwaysPhotographs','demarkationOfPathwaysRemark' ,
                    'isSignagesLabelingCheck','signagesLabelingCondition' ,'signagesLabelingPhotographs','signagesLabelingRemarks',
                    'isRegularHealthCheckup','regularHealthCheckupCondition','regularHealthCheckupPhotographs','regularHealthCheckupRemarks',
                    'isAvailabilityOfDoctor', 'availabilityOfDoctorCondition','availabilityOfDoctorPhotographs','availabilityOfDoctorRemarks',
                        'isFirstAidKit','firstAidKitCondition' ,'firstAidKitPhotographs','firstAidKitRemarks',
                    'isDrinkingWaterCheck','drinkingWaterCondition' ,'drinkingWaterPhotographs','drinkingWaterRemarks',
                        'isToilet', 'toiletCondition','toiletPhotograph','toiletRemarks',
                        'genralphotographs','documents','remarks'
            # Add more fields as needed
        )

        if not data:
            return Response({'status':'error','message':'Data Not Found'}, status=400)


        else:
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfMonitoring': 'Date of Monitoring',
                'constructionSiteName': 'Construction Site Name',
                'constructionSiteId': 'Construction Site ID',
                'isDemarkationOfPathways': 'Demarkation of Pathways Available',
                'demarkationOfPathwaysCondition': 'Pathways Demarkation Condition',
                'demarkationOfPathwaysPhotographs': 'Pathways Demarkation Photographs',
                'demarkationOfPathwaysRemark': 'Pathways Demarkation Remarks',
                'isSignagesLabelingCheck': 'Signages and Labeling Available',
                'signagesLabelingCondition': 'Signages and Labeling Condition',
                'signagesLabelingPhotographs': 'Signages and Labeling Photographs',
                'signagesLabelingRemarks': 'Signages and Labeling Remarks',
                'isRegularHealthCheckup': 'Regular Health Checkup Available',
                'regularHealthCheckupCondition': 'Health Checkup Condition',
                'regularHealthCheckupPhotographs': 'Health Checkup Photographs',
                'regularHealthCheckupRemarks': 'Health Checkup Remarks',
                'isAvailabilityOfDoctor': 'Doctor Availability',
                'availabilityOfDoctorCondition': 'Doctor Availability Condition',
                'availabilityOfDoctorPhotographs': 'Doctor Availability Photographs',
                'availabilityOfDoctorRemarks': 'Doctor Availability Remarks',
                'isFirstAidKit': 'First Aid Kit Available',
                'firstAidKitCondition': 'First Aid Kit Condition',
                'firstAidKitPhotographs': 'First Aid Kit Photographs',
                'firstAidKitRemarks': 'First Aid Kit Remarks',
                'isDrinkingWaterCheck': 'Drinking Water Available',
                'drinkingWaterCondition': 'Drinking Water Condition',
                'drinkingWaterPhotographs': 'Drinking Water Photographs',
                'drinkingWaterRemarks': 'Drinking Water Remarks',
                'isToilet': 'Toilet Available',
                'toiletCondition': 'Toilet Condition',
                'toiletPhotograph': 'Toilet Photographs',
                'toiletRemarks': 'Toilet Remarks',
                'genralphotographs': 'General Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=ConstructionCampReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class ConstructionCampFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')

    class Meta:
        model = ConstructionSiteDetails
        fields = ['quarter', 'year', 'constructionSiteName']



class ConstructionCampReportQuaterExcelDownload(generics.ListAPIView):
    serializer_class = ConstructionCampReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = ConstructionSiteDetails.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values(
                    'id','quarter', 'packages','dateOfMonitoring' ,'constructionSiteName' , 'constructionSiteId',
                    'isDemarkationOfPathways','demarkationOfPathwaysCondition','demarkationOfPathwaysPhotographs','demarkationOfPathwaysRemark' ,
                    'isSignagesLabelingCheck','signagesLabelingCondition' ,'signagesLabelingPhotographs','signagesLabelingRemarks',
                    'isRegularHealthCheckup','regularHealthCheckupCondition','regularHealthCheckupPhotographs','regularHealthCheckupRemarks',
                    'isAvailabilityOfDoctor', 'availabilityOfDoctorCondition','availabilityOfDoctorPhotographs','availabilityOfDoctorRemarks',
                        'isFirstAidKit','firstAidKitCondition' ,'firstAidKitPhotographs','firstAidKitRemarks',
                    'isDrinkingWaterCheck','drinkingWaterCondition' ,'drinkingWaterPhotographs','drinkingWaterRemarks',
                        'isToilet', 'toiletCondition','toiletPhotograph','toiletRemarks',
                        'genralphotographs','documents','remarks'
            # Add more fields as needed
        )
        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfMonitoring': 'Date of Monitoring',
                'constructionSiteName': 'Construction Site Name',
                'constructionSiteId': 'Construction Site ID',
                'isDemarkationOfPathways': 'Demarkation of Pathways Available',
                'demarkationOfPathwaysCondition': 'Pathways Demarkation Condition',
                'demarkationOfPathwaysPhotographs': 'Pathways Demarkation Photographs',
                'demarkationOfPathwaysRemark': 'Pathways Demarkation Remarks',
                'isSignagesLabelingCheck': 'Signages and Labeling Available',
                'signagesLabelingCondition': 'Signages and Labeling Condition',
                'signagesLabelingPhotographs': 'Signages and Labeling Photographs',
                'signagesLabelingRemarks': 'Signages and Labeling Remarks',
                'isRegularHealthCheckup': 'Regular Health Checkup Available',
                'regularHealthCheckupCondition': 'Health Checkup Condition',
                'regularHealthCheckupPhotographs': 'Health Checkup Photographs',
                'regularHealthCheckupRemarks': 'Health Checkup Remarks',
                'isAvailabilityOfDoctor': 'Doctor Availability',
                'availabilityOfDoctorCondition': 'Doctor Availability Condition',
                'availabilityOfDoctorPhotographs': 'Doctor Availability Photographs',
                'availabilityOfDoctorRemarks': 'Doctor Availability Remarks',
                'isFirstAidKit': 'First Aid Kit Available',
                'firstAidKitCondition': 'First Aid Kit Condition',
                'firstAidKitPhotographs': 'First Aid Kit Photographs',
                'firstAidKitRemarks': 'First Aid Kit Remarks',
                'isDrinkingWaterCheck': 'Drinking Water Available',
                'drinkingWaterCondition': 'Drinking Water Condition',
                'drinkingWaterPhotographs': 'Drinking Water Photographs',
                'drinkingWaterRemarks': 'Drinking Water Remarks',
                'isToilet': 'Toilet Available',
                'toiletCondition': 'Toilet Condition',
                'toiletPhotograph': 'Toilet Photographs',
                'toiletRemarks': 'Toilet Remarks',
                'genralphotographs': 'General Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=ConstructionCampReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response




# -------------------------------------------------------------------------

class PAPReportPackageView(ListAPIView):
    serializer_class = PAPReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, *args, **kwargs):
        try:
            data = PAP.objects.filter(packages=packages).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)
            pap_data = PAPReportSerializer(data, many=True).data

            for feature in pap_data['features']:
                feature['properties']['id'] = feature['id']
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "PAP": pap_data},
                            status=status.HTTP_200_OK)
        except Exception:
            return Response({'Message': 'There is no data available for the Package or Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)



class PAPReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = PAPReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = PAP.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values(
            'id','quarter', 'packages','dateOfMonitoring','dateOfIdentification','PAPID','firstName','middleName','lastName', 'cadastralMapID', 'cadastralMapDocuments', 
                  'addressLine1','streetName','pincode','eligibility', 'categoryOfPap','dateOfIdentification',
                  'areaOfAsset','legalStatus','legalDocuments', 
                   'actionTaken', 'notAgreedReason','remarks', 'presentPhotograph', 'documents'
            # Add more fields as needed
        )
        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfMonitoring': 'Monitoring Date',
                'dateOfIdentification': 'Date of Identification',
                'PAPID': 'PAP ID',
                'firstName': 'First Name',
                'middleName': 'Middle Name',
                'lastName': 'Last Name',
                'cadastralMapID': 'Cadastral Map ID',
                'cadastralMapDocuments': 'Cadastral Map Documents',
                'addressLine1': 'Address Line 1',
                'streetName': 'Street Name',
                'pincode': 'Pincode',
                'eligibility': 'Eligibility Status',
                'categoryOfPap': 'Category of PAP',
                'areaOfAsset': 'Area of Asset',
                'legalStatus': 'Legal Status',
                'legalDocuments': 'Legal Documents',
                'actionTaken': 'Action Taken',
                'notAgreedReason': 'Reason for Non-Agreement',
                'remarks': 'Remarks',
                'presentPhotograph': 'Photograph',
                'documents': 'Documents'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=PAPReportPackage.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response



class PAPReportQuarterView(ListAPIView):
    serializer_class = PAPReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, quarter, year):
        try:
            data = PAP.objects.filter(quarter=quarter, dateOfMonitoring__year=year).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)
            papdata = self.serializer_class(data, many=True).data

            for feature in papdata['features']:
                feature['properties']['id'] = feature['id'] 

            
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "PAP": papdata},
                            status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available for the Package or Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)




class PAPReportExcelQuaterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')

    class Meta:
        model = PAP
        fields = ['quarter', 'year']



class PAPReportExcelQuaterExcelDownload(generics.ListAPIView):
    serializer_class = PAPReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = PAPReportExcelQuaterFilter

    def get_queryset(self):
        queryset = PAP.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values(
            'id','quarter', 'packages','dateOfMonitoring','dateOfIdentification','PAPID','firstName','middleName','lastName', 'cadastralMapID', 'cadastralMapDocuments', 
                  'addressLine1','streetName','pincode','eligibility', 'categoryOfPap','dateOfIdentification',
                  'areaOfAsset','legalStatus','legalDocuments', 
                   'actionTaken', 'notAgreedReason','remarks', 'presentPhotograph', 'documents'
            # Add more fields as needed
        )


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfMonitoring': 'Monitoring Date',
                'dateOfIdentification': 'Date of Identification',
                'PAPID': 'PAP ID',
                'firstName': 'First Name',
                'middleName': 'Middle Name',
                'lastName': 'Last Name',
                'cadastralMapID': 'Cadastral Map ID',
                'cadastralMapDocuments': 'Cadastral Map Documents',
                'addressLine1': 'Address Line 1',
                'streetName': 'Street Name',
                'pincode': 'Pincode',
                'eligibility': 'Eligibility Status',
                'categoryOfPap': 'Category of PAP',
                'areaOfAsset': 'Area of Asset',
                'legalStatus': 'Legal Status',
                'legalDocuments': 'Legal Documents',
                'actionTaken': 'Action Taken',
                'notAgreedReason': 'Reason for Non-Agreement',
                'remarks': 'Remarks',
                'presentPhotograph': 'Photograph',
                'documents': 'Documents'
            })


            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=PAPReportExcelQuater.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


# ------------------------------------------------------------
class RehabilitationReportPackageView(ListAPIView):
    serializer_class = RehabilitationReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, *args, **kwargs):
        try:
            data = Rehabilitation.objects.filter(packages=packages).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',},  status=status.HTTP_400_BAD_REQUEST)
            rehabilitation_data = RehabilitationReportSerializer(data, many=True).data
            for feature in rehabilitation_data['features']:
                feature['properties']['id'] = feature['id']
    
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "Rehabilated_PAP_Package": rehabilitation_data},
                            status=status.HTTP_200_OK)
        except Exception:
            return Response({'Message': 'There is no data available for the Package or Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)





class RehabilitationReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = RehabilitationReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = Rehabilitation.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('quarter','packages','dateOfRehabilitation' ,'PAPID',
                   'firstName', 'middleName', 'lastName', 'compensationStatus', 'agreedUpon', 'processStatus',
                   'cashCompensationAmount',
                   'typeOfCompensation', 'otherCompensationType' ,
                   'addressLine1','streetName','pincode',
                   'rehabLocation', 'allowance', 'area',
                   'isShiftingAllowance','shiftingAllowanceAmount',
                   'isLivelihoodSupport', 'livelihoodSupportAmount',
                   'isTraining','trainingRemarks', 'typeOfStructure',
                   'isRelocationAllowance' ,'RelocationAllowanceAmount' ,'isfinancialSupport',
                   'financialSupportAmount','isCommunityEngagement','isEngagementType',
                   'photographs' , 'documents','remarks')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)

        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
            # Rename columns to more appropriate names
            df = df.rename(columns={
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfRehabilitation': 'Rehabilitation Date',
                'PAPID': 'PAP ID',
                'firstName': 'First Name',
                'middleName': 'Middle Name',
                'lastName': 'Last Name',
                'compensationStatus': 'Compensation Status',
                'agreedUpon': 'Agreed Upon',
                'processStatus': 'Process Status',
                'cashCompensationAmount': 'Cash Compensation Amount',
                'typeOfCompensation': 'Type of Compensation',
                'otherCompensationType': 'Other Compensation Type',
                'addressLine1': 'Address Line 1',
                'streetName': 'Street Name',
                'pincode': 'Pincode',
                'rehabLocation': 'Rehabilitation Location',
                'allowance': 'Allowance',
                'area': 'Area (sq.m.)',
                'isShiftingAllowance': 'Shifting Allowance',
                'shiftingAllowanceAmount': 'Shifting Allowance Amount',
                'isLivelihoodSupport': 'Livelihood Support',
                'livelihoodSupportAmount': 'Livelihood Support Amount',
                'isTraining': 'Training',
                'trainingRemarks': 'Training Remarks',
                'typeOfStructure': 'Type of Structure',
                'isRelocationAllowance': 'Relocation Allowance',
                'RelocationAllowanceAmount': 'Relocation Allowance Amount',
                'isfinancialSupport': 'Financial Support',
                'financialSupportAmount': 'Financial Support Amount',
                'isCommunityEngagement': 'Community Engagement Involved',
                'isEngagementType': 'Type of Engagement',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=RehabilitationReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response




class RehabilitationReportQuarterView(ListAPIView):
    serializer_class = RehabilitationReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, quarter, year, *args, **kwargs):
        try:
            data = Rehabilitation.objects.filter(quarter=quarter, dateOfRehabilitation__year=year).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)
            Rehabilitation_data = self.serializer_class(data, many=True).data

            for feature in Rehabilitation_data['features']:
                feature['properties']['id'] = feature['id'] 

            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "Rehabilated_PAP_Quarter_wise": Rehabilitation_data},
                            status=status.HTTP_200_OK)
        except Exception:
            return Response({'Message': 'There is no data available for the Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)



class RehabilitationReportQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfRehabilitation__year', label='Year')

    class Meta:
        model = Rehabilitation
        fields = ['quarter', 'year']



class RehabilitationReportQuarterExcelDownload(generics.ListAPIView):
    serializer_class = RehabilitationReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = RehabilitationReportQuarterFilter

    def get_queryset(self):
        queryset = Rehabilitation.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('quarter','packages','dateOfRehabilitation' ,'PAPID',
                   'firstName', 'middleName', 'lastName', 'compensationStatus', 'agreedUpon', 'processStatus',
                   'cashCompensationAmount',
                   'typeOfCompensation', 'otherCompensationType' ,
                   'addressLine1','streetName','pincode',
                   'rehabLocation', 'allowance', 'area',
                   'isShiftingAllowance','shiftingAllowanceAmount',
                   'isLivelihoodSupport', 'livelihoodSupportAmount',
                   'isTraining','trainingRemarks', 'typeOfStructure',
                   'isRelocationAllowance' ,'RelocationAllowanceAmount' ,'isfinancialSupport',
                   'financialSupportAmount','isCommunityEngagement','isEngagementType',
                   'photographs' , 'documents','remarks')

        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)

        else:
                
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
             # Rename columns to more appropriate names
            df = df.rename(columns={
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfRehabilitation': 'Rehabilitation Date',
                'PAPID': 'PAP ID',
                'firstName': 'First Name',
                'middleName': 'Middle Name',
                'lastName': 'Last Name',
                'compensationStatus': 'Compensation Status',
                'agreedUpon': 'Agreed Upon',
                'processStatus': 'Process Status',
                'cashCompensationAmount': 'Cash Compensation Amount',
                'typeOfCompensation': 'Type of Compensation',
                'otherCompensationType': 'Other Compensation Type',
                'addressLine1': 'Address Line 1',
                'streetName': 'Street Name',
                'pincode': 'Pincode',
                'rehabLocation': 'Rehabilitation Location',
                'allowance': 'Allowance',
                'area': 'Area (sq.m.)',
                'isShiftingAllowance': 'Shifting Allowance',
                'shiftingAllowanceAmount': 'Shifting Allowance Amount',
                'isLivelihoodSupport': 'Livelihood Support',
                'livelihoodSupportAmount': 'Livelihood Support Amount',
                'isTraining': 'Training',
                'trainingRemarks': 'Training Remarks',
                'typeOfStructure': 'Type of Structure',
                'isRelocationAllowance': 'Relocation Allowance',
                'RelocationAllowanceAmount': 'Relocation Allowance Amount',
                'isfinancialSupport': 'Financial Support',
                'financialSupportAmount': 'Financial Support Amount',
                'isCommunityEngagement': 'Community Engagement Involved',
                'isEngagementType': 'Type of Engagement',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=RehabilitationReportQuater.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response




'''----------------------- Env Monitoring Report View------------------------------'''


class AirReportPackageView(ListAPIView):
    serializer_class = AirReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, *args, **kwargs):
        try:
            data = Air.objects.filter(packages=packages).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)
            air_data = AirReportSerializer(data, many=True).data

            for feature in air_data['features']:
                feature['properties']['id'] = feature['id']   
                
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "Air_data": air_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available for the Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)




class AirReportReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = AirReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = Air.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','packages','month','dateOfMonitoring','PM10','PM2_5',
                 'SO2','NOx','CO','AQI','place_location')

        print(data)

        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)

        else:
            

            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns as needed
            df = df.rename(columns={
                'id': 'ID',
                'dateOfMonitoring': 'Monitoring Date',
                'quarter': 'Quarter',
                'packages': 'Packages',
                'month': 'Month',
                'PM10': 'PM10 Level',
                'PM2_5': 'PM2.5 Level',
                'SO2': 'SO2 Level',
                'NOx': 'NOx Level',
                'CO': 'CO Level',
                'AQI': 'Air Quality Index',
                'place_location': 'Location'
            })
            
            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=AirReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response





class AirReportQuarterView(ListAPIView):
    serializer_class = AirReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, month, year, *args, **kwargs):
        try:
            data = Air.objects.filter(month=month, dateOfMonitoring__year=year).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)

            airdata = self.serializer_class(data, many=True).data

            for feature in airdata['features']:
                feature['properties']['id'] = feature['id'] 
                
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "Air_data": airdata},
                            status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available for the Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)



class AirReportQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')
    month = django_filters.CharFilter(method='filter_by_month', label='Month')

    def filter_by_month(self, queryset, name, value):
        # if value.isdigit():
        #     # If the value is a digit, return the queryset as is
        #     return queryset

        try:
            month_number = list(calendar.month_name).index(value.capitalize())
            return queryset.filter(dateOfMonitoring__month=month_number)
        except ValueError:
            # If the value is not a valid month name, return an empty queryset
            return queryset.none()
    class Meta:
        model = Air
        fields = [ 'month','year']





class AirReportQuarterExcelDownload(generics.ListAPIView):
    serializer_class = AirReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = AirReportQuarterFilter

    def get_queryset(self):
        queryset = Air.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','packages','month','dateOfMonitoring','PM10','PM2_5',
                 'SO2','NOx','CO','AQI', 'place_location')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns as needed
            df = df.rename(columns={
                'id': 'ID',
                'dateOfMonitoring': 'Monitoring Date',
                'quarter': 'Quarter',
                'packages': 'Packages',
                'month': 'Month',
                'PM10': 'PM10 Level',
                'PM2_5': 'PM2.5 Level',
                'SO2': 'SO2 Level',
                'NOx': 'NOx Level',
                'CO': 'CO Level',
                'AQI': 'Air Quality Index',
                'place_location': 'Location'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=AirReportQuater.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class NoiseReportpackageView(ListAPIView):
    serializer_class = NoiseReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, *args, **kwargs):
        try:
            data = Noise.objects.filter(packages=packages).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)

            noise_data = NoiseReportSerializer(data, many=True).data
            for feature in noise_data['features']:
                feature['properties']['id'] = feature['id']
                
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "Noise_data": noise_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available for the Package',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)



class NoiseReportReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = NoiseReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = Noise.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id', 'quarter','month','packages','dateOfMonitoringThree' , 'noiseLevel_day', 'noiseLevel_night', 'monitoringPeriod_day', 'monitoringPeriod_night', 'typeOfArea', 'isWithinLimit_day', 'isWithinLimit_night' ,'place_location' )

        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
                
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns with more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'dateOfMonitoringThree': 'Monitoring Date',
                'quarter': 'Quarter',
                'month': 'Month',
                'packages': 'Packages',
                'noiseLevel_day': 'Day Noise Level (dB)',
                'noiseLevel_night': 'Night Noise Level (dB)',
                'monitoringPeriod_day': 'Day Monitoring Period',
                'monitoringPeriod_night': 'Night Monitoring Period',
                'typeOfArea': 'Area Type',
                'isWithinLimit_day': 'Within Limit (Day)',
                'isWithinLimit_night': 'Within Limit (Night)',
                'place_location': 'Location'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=NoiseReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class NoiseReportQuarterView(ListAPIView):
    serializer_class = NoiseReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, month, year, *args, **kwargs):
        try:
            data = Noise.objects.filter(month=month, dateOfMonitoringThree__year=year).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)

            Noise_data = self.serializer_class(data, many=True).data

            for feature in Noise_data['features']:
                feature['properties']['id'] = feature['id'] 
                
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "Noise_data": Noise_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available for the Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)





class NoiseReportQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoringThree__year', label='Year')
    month = django_filters.CharFilter(method='filter_by_month', label='Month')

    def filter_by_month(self, queryset, name, value):
        # if value.isdigit():
        #     # If the value is a digit, return the queryset as is
        #     return queryset

        try:
            month_number = list(calendar.month_name).index(value.capitalize())
            return queryset.filter(dateOfMonitoringThree__month=month_number)
        except ValueError:
            # If the value is not a valid month name, return an empty queryset
            return queryset.none()
    class Meta:
        model = Noise
        fields = ['month', 'year']



class NoiseReportQuarterExcelDownload(generics.ListAPIView):
    serializer_class = NoiseReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = NoiseReportQuarterFilter

    def get_queryset(self):
        queryset = Noise.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id', 'quarter','month','packages','dateOfMonitoringThree' , 'noiseLevel_day', 'noiseLevel_night', 'monitoringPeriod_day', 'monitoringPeriod_night', 'typeOfArea', 'isWithinLimit_day', 'isWithinLimit_night','place_location' )

        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        # Create a Pandas DataFrame
        else:
            
            df = pd.DataFrame(data)
           
            df = df.rename(columns={
                'id': 'ID',
                'dateOfMonitoringThree': 'Monitoring Date',
                'quarter': 'Quarter',
                'month': 'Month',
                'packages': 'Packages',
                'noiseLevel_day': 'Day Noise Level (dB)',
                'noiseLevel_night': 'Night Noise Level (dB)',
                'monitoringPeriod_day': 'Day Monitoring Period',
                'monitoringPeriod_night': 'Night Monitoring Period',
                'typeOfArea': 'Area Type',
                'isWithinLimit_day': 'Within Limit (Day)',
                'isWithinLimit_night': 'Within Limit (Night)',
                'place_location': 'Location'
            })
            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=NoiseReportQuater.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response

class waterReportPackageView(ListAPIView):
    serializer_class = waterReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, *args, **kwargs):
        try:
            data = water.objects.filter(packages=packages).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)

            water_data = waterReportSerializer(data, many=True).data
            for feature in water_data['features']:
                feature['properties']['id'] = feature['id']   
                
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "water_data": water_data}, status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available for the Package',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)


class waterReportReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = waterExcelReportSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = water.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','packages','month', 'dateOfMonitoringTwo','qualityOfWater' , 'sourceOfWater' , 'trueColor', 'turbidity', 'odour', 'waterDisposal', 'WQI', 'pH', 'totalHardnessAsCaCO3', 'calcium', 'totalAlkalinityAsCaCO3', 'chlorides', 'magnesium', 'totalDissolvedSolids', 'sulphate', 'nitrate', 'fluoride', 'iron', 'zinc', 'copper', 'aluminum', 'nickel', 'manganese', 'phenolicCompounds', 'sulphide', 'cadmium', 'cyanide', 'lead', 'mercury', 'totalArsenic', 'totalChromium','place_location' )


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)

        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'packages': 'Packages',
                'month': 'Month',
                'dateOfMonitoringTwo': 'Monitoring Date',
                'qualityOfWater': 'Water Quality',
                'sourceOfWater': 'Water Source',
                'trueColor': 'True Color (CU)',
                'turbidity': 'Turbidity (NTU)',
                'odour': 'Odour',
                'waterDisposal': 'Water Disposal Method',
                'WQI': 'Water Quality Index (WQI)',
                'pH': 'pH Level',
                'totalHardnessAsCaCO3': 'Total Hardness (as CaCO3)',
                'calcium': 'Calcium (mg/L)',
                'totalAlkalinityAsCaCO3': 'Total Alkalinity (as CaCO3)',
                'chlorides': 'Chlorides (mg/L)',
                'magnesium': 'Magnesium (mg/L)',
                'totalDissolvedSolids': 'Total Dissolved Solids (TDS)',
                'sulphate': 'Sulphate (mg/L)',
                'nitrate': 'Nitrate (mg/L)',
                'fluoride': 'Fluoride (mg/L)',
                'iron': 'Iron (mg/L)',
                'zinc': 'Zinc (mg/L)',
                'copper': 'Copper (mg/L)',
                'aluminum': 'Aluminum (mg/L)',
                'nickel': 'Nickel (mg/L)',
                'manganese': 'Manganese (mg/L)',
                'phenolicCompounds': 'Phenolic Compounds (mg/L)',
                'sulphide': 'Sulphide (mg/L)',
                'cadmium': 'Cadmium (mg/L)',
                'cyanide': 'Cyanide (mg/L)',
                'lead': 'Lead (mg/L)',
                'mercury': 'Mercury (mg/L)',
                'totalArsenic': 'Total Arsenic (mg/L)',
                'totalChromium': 'Total Chromium (mg/L)',
                'place_location': 'Location'
            })
            
            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=WaterReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class waterReportQuarterView(ListAPIView):
    serializer_class = waterReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, month, year, *args, **kwargs):
        # try:
            data = water.objects.filter(month=month, dateOfMonitoringTwo__year=year).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 'status' : 'success'},  status=status.HTTP_400_BAD_REQUEST)

            water_data = self.serializer_class(data, many=True).data

            for feature in water_data['features']:
                feature['properties']['id'] = feature['id'] 
                
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "water_data": water_data},
                            status=status.HTTP_200_OK)
        # except:
        #     return Response({'Message': 'There is no data available for the Quarter',
        #                     'status' : 'Failed'},
        #                     status=status.HTTP_400_BAD_REQUEST)



class WaterReportQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoringTwo__year', label='Year')
    month = django_filters.CharFilter(method='filter_by_month', label='Month')

    def filter_by_month(self, queryset, name, value):
        # if value.isdigit():
        #     # If the value is a digit, return the queryset as is
        #     return queryset

        try:
            month_number = list(calendar.month_name).index(value.capitalize())
            return queryset.filter(dateOfMonitoringTwo__month=month_number)
        except ValueError:
            # If the value is not a valid month name, return an empty queryset
            return queryset.none()
    class Meta:
        model = water
        fields = ['month', 'year']



class WaterReportQuarterExcelDownload(generics.ListAPIView):
    serializer_class = WaterReportQuarterFilter
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = WaterReportQuarterFilter

    def get_queryset(self):
        queryset = water.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','packages','month', 'dateOfMonitoringTwo','qualityOfWater' , 'sourceOfWater' , 'trueColor', 'turbidity', 'odour', 'waterDisposal', 'WQI', 'pH', 'totalHardnessAsCaCO3', 'calcium', 'totalAlkalinityAsCaCO3', 'chlorides', 'magnesium', 'totalDissolvedSolids', 'sulphate', 'nitrate', 'fluoride', 'iron', 'zinc', 'copper', 'aluminum', 'nickel', 'manganese', 'phenolicCompounds', 'sulphide', 'cadmium', 'cyanide', 'lead', 'mercury', 'totalArsenic', 'totalChromium','place_location' )

        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
                
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'packages': 'Packages',
                'month': 'Month',
                'dateOfMonitoringTwo': 'Monitoring Date',
                'qualityOfWater': 'Water Quality',
                'sourceOfWater': 'Water Source',
                'trueColor': 'True Color (CU)',
                'turbidity': 'Turbidity (NTU)',
                'odour': 'Odour',
                'waterDisposal': 'Water Disposal Method',
                'WQI': 'Water Quality Index (WQI)',
                'pH': 'pH Level',
                'totalHardnessAsCaCO3': 'Total Hardness (as CaCO3)',
                'calcium': 'Calcium (mg/L)',
                'totalAlkalinityAsCaCO3': 'Total Alkalinity (as CaCO3)',
                'chlorides': 'Chlorides (mg/L)',
                'magnesium': 'Magnesium (mg/L)',
                'totalDissolvedSolids': 'Total Dissolved Solids (TDS)',
                'sulphate': 'Sulphate (mg/L)',
                'nitrate': 'Nitrate (mg/L)',
                'fluoride': 'Fluoride (mg/L)',
                'iron': 'Iron (mg/L)',
                'zinc': 'Zinc (mg/L)',
                'copper': 'Copper (mg/L)',
                'aluminum': 'Aluminum (mg/L)',
                'nickel': 'Nickel (mg/L)',
                'manganese': 'Manganese (mg/L)',
                'phenolicCompounds': 'Phenolic Compounds (mg/L)',
                'sulphide': 'Sulphide (mg/L)',
                'cadmium': 'Cadmium (mg/L)',
                'cyanide': 'Cyanide (mg/L)',
                'lead': 'Lead (mg/L)',
                'mercury': 'Mercury (mg/L)',
                'totalArsenic': 'Total Arsenic (mg/L)',
                'totalChromium': 'Total Chromium (mg/L)',
                'place_location': 'Location'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=WaterReportQuater.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


# Existing Tree
class TreeMangementReportPackage(ListAPIView):
    serializer_class = treeManagementSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, *args, **kwargs):
        try:
            data = ExistingTreeManagment.objects.filter(packages=packages).order_by('-id')
            if not data.exists():
                return Response({'message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)

            existing_tree_data = treeManagementSerializer(data, many=True).data 
            for feature in existing_tree_data['features']:
                feature['properties']['id'] = feature['id']
            
            return Response({'message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            'existing_tree_data': existing_tree_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'message': 'There is no data available for the Package',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)


class TreeManagementReportQuarterView(ListAPIView):
    serializer_class = treeManagementSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, quarter, year, *args, **kwargs):
        try:
            data = ExistingTreeManagment.objects.filter(quarter=quarter, dateOfMonitoring__year=year).order_by('-id')
            if not data.exists():
                return Response({'message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)

            existing_tree_data = treeManagementSerializer(data, many=True).data
            for feature in existing_tree_data['features']:
                feature['properties']['id'] = feature['id']
                
            return Response({'message': 'data Fetched Successfully',
                             'status' : 'success' , 
                             'existing_tree_data': existing_tree_data},
                            status=200)
        except:
            return Response({'message': 'There is no data available for the Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)


class treeManagementReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = treeManagementExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = ExistingTreeManagment.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','month','dateOfMonitoring','packages','treeID','commanName' ,'botanicalName',
                    'condition', 'actionTaken', 'photographs', 'documents','remarks')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)

        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
            # Rename columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'month': 'Month',
                'dateOfMonitoring': 'Monitoring Date',
                'packages': 'Package',
                'treeID': 'Tree ID',
                'commanName': 'Common Name',
                'botanicalName': 'Botanical Name',
                'condition': 'Tree Condition',
                'actionTaken': 'Action Taken',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=ExistingTreeManagmentReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class ExistingTreeManagmentQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')
    quarter = django_filters.CharFilter(field_name='quarter', label='quarter')

    class Meta:
        model = ExistingTreeManagment
        fields = ['quarter', 'year']



class TreeManagementQuarterExcelDownload(generics.ListAPIView):
    serializer_class = treeManagementExcelSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = ExistingTreeManagmentQuarterFilter

    def get_queryset(self):
        queryset = ExistingTreeManagment.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','month','dateOfMonitoring','packages','treeID','commanName' ,'botanicalName','place_location',
                    'condition', 'actionTaken', 'photographs', 'documents','remarks')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
            # Rename columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'month': 'Month',
                'dateOfMonitoring': 'Monitoring Date',
                'packages': 'Package',
                'treeID': 'Tree ID',
                'commanName': 'Common Name',
                'botanicalName': 'Botanical Name',
                'condition': 'Tree Condition',
                'actionTaken': 'Action Taken',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=TreeManegmanetQuater.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


# New Tree
class NewTreeReportPackage(ListAPIView):
    serializer_class = NewTreeManagementSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, *args, **kwargs):
        try:
            data = NewTreeManagement.objects.filter(packages=packages).order_by('-id')
            print(data)
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)

            new_tree_data = NewTreeManagementSerializer(data, many=True).data
            print(new_tree_data)
            # new_tree_data = self.get_serializer(data, many=True).data
            for feature in new_tree_data['features']:
                feature['properties']['id'] = feature['id']

            # for feature in new_tree_data['new_tree_data']['features']:
            #     feature['properties']['id'] = feature['id']
            
            return Response({'message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            'new_tree_data': new_tree_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available for the Package',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)


# New Tree
class NewTreeReportQuarterView(ListAPIView):
    serializer_class = NewTreeManagementSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, quarter, year, *args, **kwargs):
        try:
            data = NewTreeManagement.objects.filter(quarter=quarter, dateOfMonitoring__year=year).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)

            # new_tree_data = self.get_serializer(data, many=True).data
            new_tree_data = treeManagementSerializer(data, many=True).data
            
            for feature in new_tree_data['features']:
                feature['properties']['id'] = feature['id']
                
            return Response({'message': 'Data Fetched Successfully',
                             'status' : 'Success' , 
                             'new_tree_data': new_tree_data},
                            status=200)
        except:
            return Response({'Message': 'There is no data available for the Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)


class NewTreePackageExcelDownload(generics.ListAPIView):
    serializer_class = ExcelNewTreeQuarterSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    

    def get_queryset(self):
        queryset = NewTreeManagement.objects.all()
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('tree', 'quarter', 'month', 'dateOfMonitoring', 'packages',
                  'commanName','place_location', 'botanicalName', 'condition', 'photographs', 'documents', 'remarks')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'month': 'Month',
                'dateOfMonitoring': 'Monitoring Date',
                'packages': 'Package',
                'treeID': 'Tree ID',
                'commanName': 'Common Name',
                'botanicalName': 'Botanical Name',
                'condition': 'Tree Condition',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Get the package filter from the request
            package = request.GET.get('packages', 'all_packages')

            # Create a response with the appropriate content type
            filename = f'New_Tree_{package}.xlsx'
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class NewTreeQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')

    class Meta:
        model = NewTreeManagement
        fields = ['quarter', 'year']



class NewTreeQuarterExcelDownload(generics.ListAPIView):
    serializer_class = ExcelOccupationalHealthQuarterSeialzier
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = NewTreeQuarterFilter

    def get_queryset(self):
        queryset = NewTreeManagement.objects.all()
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('tree', 'quarter', 'month', 'dateOfMonitoring', 'packages',
                  'commanName', 'botanicalName', 'condition', 'photographs', 'documents', 'remarks''place_location')
        
        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'month': 'Month',
                'dateOfMonitoring': 'Monitoring Date',
                'packages': 'Package',
                'treeID': 'Tree ID',
                'commanName': 'Common Name',
                'botanicalName': 'Botanical Name',
                'condition': 'Tree Condition',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })


            # Get the quarter and year filter from the request
            quarter = request.GET.get('quarter', 'all_quarters')
            year = request.GET.get('year', 'all_years')

            # Create a response with the appropriate content type
            filename = f'New_Tree_{year}_{quarter}.xlsx'
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class WasteTreatmentsPackageView(ListAPIView):
    serializer_class = wasteTreatmentsSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, *args, **kwargs):
        try:
            data = WasteTreatments.objects.filter(packages=packages).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',},  status=status.HTTP_400_BAD_REQUEST)
            waste_data = wasteTreatmentsSerializer(data, many=True).data
            for feature in waste_data['features']:
                feature['properties']['id'] = feature['id']   

            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "wasteManagementdata": waste_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available for the Package',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)


class WasteTreatmentsQuarterView(ListAPIView):
    serializer_class = wasteTreatmentsSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, quarter, year, *args, **kwargs):
        try:
            data = WasteTreatments.objects.filter(quarter=quarter, dateOfMonitoring__year=year).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)

            Waste_data = self.serializer_class(data, many=True).data

            for feature in Waste_data['features']:
                feature['properties']['id'] = feature['id'] 
                
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                             "wasteManagementData": Waste_data}, status=200)

        except:
            return Response({'Message': 'There is no data available for the Quarter',
                            'status' : 'Failed'} , status=status.HTTP_400_BAD_REQUEST)


class wasteTreatmentReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = wasteTreatmentsExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = WasteTreatments.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','month','packages','dateOfMonitoring', 'GISPermitsTransportationDocuments', 'TransportationVechicalHasPermissionDocuments', 'wasteOilQnt', 'CCPCPaintSludgeQnt', 'filterQnt', 'airFiltersQnt', 'usedCartridgesQnt', 'plasticQnt', 'paperQnt', 'woodQnt', 'bottlesQnt', 'rubberQnt', 'bioDegradableQuantity', 'bioMedicalQuantity', 'metalScrapeQuantity', 'eWasteQuantity', 'constructionWasteQuantity', 'waste_handlingLocation', 'photographs' , 'documents','remarks')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more descriptive names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'month': 'Month',
                'packages': 'Packages',
                'dateOfMonitoring': 'Monitoring Date',
                'GISPermitsTransportationDocuments': 'GIS Permits & Transportation Documents',
                'TransportationVechicalHasPermissionDocuments': 'Vehicle Permission Documents',
                'wasteOilQnt': 'Waste Oil Quantity (Liters)',
                'CCPCPaintSludgeQnt': 'CCPC Paint Sludge Quantity (kg)',
                'filterQnt': 'Filter Quantity (kg)',
                'airFiltersQnt': 'Air Filters Quantity (kg)',
                'usedCartridgesQnt': 'Used Cartridges Quantity (kg)',
                'plasticQnt': 'Plastic Quantity (kg)',
                'paperQnt': 'Paper Quantity (kg)',
                'woodQnt': 'Wood Quantity (kg)',
                'bottlesQnt': 'Bottles Quantity (kg)',
                'rubberQnt': 'Rubber Quantity (kg)',
                'bioDegradableQuantity': 'Bio-degradable Waste Quantity (kg)',
                'bioMedicalQuantity': 'Bio-medical Waste Quantity (kg)',
                'metalScrapeQuantity': 'Metal Scrap Quantity (kg)',
                'eWasteQuantity': 'E-waste Quantity (kg)',
                'constructionWasteQuantity': 'Construction Waste Quantity (kg)',
                'waste_handlingLocation': 'Waste Handling Location',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=wasteTreatmentReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class WasteTreatmentsQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')
    quarter = django_filters.CharFilter(field_name='quarter', label='quarter')
    # month = django_filters.CharFilter(method='filter_by_month', label='Month')

    # def filter_by_month(self, queryset, name, value):
    #     # if value.isdigit():
    #     #     # If the value is a digit, return the queryset as is
    #     #     return queryset

    #     try:
    #         month_number = list(calendar.month_name).index(value.capitalize())
    #         return queryset.filter(dateOfMonitoring__month=month_number)
    #     except ValueError:
    #         # If the value is not a valid month name, return an empty queryset
    #         return queryset.none()
    class Meta:
        model = WasteTreatments
        fields = ['quarter', 'year']



class wasteTreatmentQuarterExcelDownload(generics.ListAPIView):
    serializer_class = wasteTreatmentsExcelSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = WasteTreatmentsQuarterFilter

    def get_queryset(self):
        queryset = WasteTreatments.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','month','packages','dateOfMonitoring', 'GISPermitsTransportationDocuments', 'TransportationVechicalHasPermissionDocuments', 'wasteOilQnt', 'CCPCPaintSludgeQnt', 'filterQnt', 'airFiltersQnt', 'usedCartridgesQnt', 'plasticQnt', 'paperQnt', 'woodQnt', 'bottlesQnt', 'rubberQnt', 'bioDegradableQuantity', 'bioMedicalQuantity', 'metalScrapeQuantity', 'eWasteQuantity', 'constructionWasteQuantity', 'waste_handlingLocation', 'photographs' , 'documents','remarks')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)

        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more descriptive names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'month': 'Month',
                'packages': 'Packages',
                'dateOfMonitoring': 'Monitoring Date',
                'GISPermitsTransportationDocuments': 'GIS Permits & Transportation Documents',
                'TransportationVechicalHasPermissionDocuments': 'Vehicle Permission Documents',
                'wasteOilQnt': 'Waste Oil Quantity (Liters)',
                'CCPCPaintSludgeQnt': 'CCPC Paint Sludge Quantity (kg)',
                'filterQnt': 'Filter Quantity (kg)',
                'airFiltersQnt': 'Air Filters Quantity (kg)',
                'usedCartridgesQnt': 'Used Cartridges Quantity (kg)',
                'plasticQnt': 'Plastic Quantity (kg)',
                'paperQnt': 'Paper Quantity (kg)',
                'woodQnt': 'Wood Quantity (kg)',
                'bottlesQnt': 'Bottles Quantity (kg)',
                'rubberQnt': 'Rubber Quantity (kg)',
                'bioDegradableQuantity': 'Bio-degradable Waste Quantity (kg)',
                'bioMedicalQuantity': 'Bio-medical Waste Quantity (kg)',
                'metalScrapeQuantity': 'Metal Scrap Quantity (kg)',
                'eWasteQuantity': 'E-waste Quantity (kg)',
                'constructionWasteQuantity': 'Construction Waste Quantity (kg)',
                'waste_handlingLocation': 'Waste Handling Location',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=WasteTreatmentQuater.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response



class MaterialManagementReporetpackageView(ListAPIView):
    serializer_class = materialManagementSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages, *args, **kwargs):
        try:
            data = MaterialManegmanet.objects.filter(packages=packages).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found','status' : 'success'},  status=status.HTTP_400_BAD_REQUEST)

            material_data = materialManagementSerializer(data, many=True).data
            for feature in material_data['features']:
                feature['properties']['id'] = feature['id']
                 
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "materialManagementData": material_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'Message': 'There is no data available for the Package',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)



class MaterialManegmanetReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = materialManagementExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = MaterialManegmanet.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        if 'packages' not in request.GET:
            return JsonResponse({'status': 'error', 'message': 'Please provide filters'}, status=400)
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','month','packages','dateOfMonitoring',
         'typeOfMaterial','source','sourceOfQuarry','materialStorageType','storageLocation',
         'materialStorageCondition','materialStoragePhotograph','approvals' ,'photographs',
          'documents','remarks')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
                
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more descriptive names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'month': 'Month',
                'packages': 'Packages',
                'dateOfMonitoring': 'Monitoring Date',
                'typeOfMaterial': 'Material Type',
                'source': 'Source',
                'sourceOfQuarry': 'Quarry Source',
                'materialStorageType': 'Storage Type',
                'storageLocation': 'Storage Location',
                'materialStorageCondition': 'Storage Condition',
                'materialStoragePhotograph': 'Storage Photograph',
                'approvals': 'Approvals',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=MaterialManagementReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response



class MaterialManagementReporetQuarterView(ListAPIView):
    serializer_class = materialManagementSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, quarter, year, *args, **kwargs):
        try:
            data = MaterialManegmanet.objects.filter(quarter=quarter, dateOfMonitoring__year=year).order_by('-id')
            if not data.exists():
                return Response({'Message': 'No data found',
                                 'status' : 'success'},  status=status.HTTP_400_BAD_REQUEST)

            Material_data = self.serializer_class(data, many=True).data

            for feature in Material_data['features']:
                feature['properties']['id'] = feature['id'] 
            return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                             "materialManagementData": Material_data},
                            status=200)
        except:
            return Response({'Message': 'There is no data available for the Quarter',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)




class materialManagementQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')
    quarter = django_filters.CharFilter(field_name='quarter', label='quarter')
    # month = django_filters.CharFilter(method='filter_by_month', label='Month')

    # def filter_by_month(self, queryset, name, value):
    #     # if value.isdigit():
    #     #     # If the value is a digit, return the queryset as is
    #     #     return queryset

    #     try:
    #         month_number = list(calendar.month_name).index(value.capitalize())
    #         return queryset.filter(dateOfMonitoring__month=month_number)
    #     except ValueError:
    #         # If the value is not a valid month name, return an empty queryset
    #         return queryset.none()
    class Meta:
        model = MaterialManegmanet
        fields = ['quarter', 'year']



class materialManagementQuarterExcelDownload(generics.ListAPIView):
    serializer_class = materialManagementExcelSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = materialManagementQuarterFilter

    def get_queryset(self):
        queryset = MaterialManegmanet.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','month','packages','dateOfMonitoring',
         'typeOfMaterial','source','sourceOfQuarry','materialStorageType','storageLocation',
         'materialStorageCondition','materialStoragePhotograph','approvals' ,'photographs',
          'documents','remarks')

        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more descriptive names
            df = df.rename(columns={
                'id': 'ID',
                'quarter': 'Quarter',
                'month': 'Month',
                'packages': 'Packages',
                'dateOfMonitoring': 'Monitoring Date',
                'typeOfMaterial': 'Material Type',
                'source': 'Source',
                'sourceOfQuarry': 'Quarry Source',
                'materialStorageType': 'Storage Type',
                'storageLocation': 'Storage Location',
                'materialStorageCondition': 'Storage Condition',
                'materialStoragePhotograph': 'Storage Photograph',
                'approvals': 'Approvals',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=MaterialManegmanetQuater.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


# geopackage serializer is not used as no location data is used, so no need to specifically add id
class PreConstructionStageComplianceReportView(ListAPIView):
    serializer_class = PreConstructionStageComplianceReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, *args, **kwargs):
        try:
            data = PreConstructionStage.objects.all().order_by('-id')
            if not data.exists():
                return Response({'message': 'Pre Construction Stage Compliance data not found','status' : '400'},  status=status.HTTP_400_BAD_REQUEST)

            pre_construction_stage_compliance_data = PreConstructionStageComplianceReportSerializer(data, many=True).data
                 
            return Response({'message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "pre_construction_stage_compliance_data": pre_construction_stage_compliance_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'message': 'There is no data available for Pre Construction Stage Compliance',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)


class PreConstructionStageComplianceExcel(generics.ListAPIView):
    # filter_backends = [DjangoFilterBackend] # discuss if we want filters of package and quarter or not

    def get_queryset(self):
        queryset = PreConstructionStage.objects.all()
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id',
            'ShiftingofUtilities', 'ResponsibilityOfShiftingofUtilities', 'CurrentStatusOfShiftingofUtilities', 'ShiftingofUtilitiesDocuments',
            'PermissionForFellingOfTrees', 'ResponsibilityOfPermissionForFellingOfTrees', 'CurrentStatusPermissionForFellingOfTrees', 'PermissionForFellingOfTreesDocuments',
            'CRZClearance', 'ResponsibilityOfCRZClearance', 'CurrentStatusCRZClearance', 'CRZClearanceDocuments',
            'ForestClearance', 'ResponsibilityOfForestClearance', 'CurrentStatusOfForestClearance', 'ForestClearanceDocuments'
        )


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'ShiftingofUtilities': 'Shifting of Utilities',
                'ResponsibilityOfShiftingofUtilities': 'Responsibility for Shifting of Utilities',
                'CurrentStatusOfShiftingofUtilities': 'Current Status of Shifting of Utilities',
                'ShiftingofUtilitiesDocuments': 'Shifting of Utilities Documents',
                'PermissionForFellingOfTrees': 'Permission for Felling of Trees',
                'ResponsibilityOfPermissionForFellingOfTrees': 'Responsibility for Felling of Trees',
                'CurrentStatusPermissionForFellingOfTrees': 'Current Status of Felling of Trees Permission',
                'PermissionForFellingOfTreesDocuments': 'Felling of Trees Documents',
                'CRZClearance': 'CRZ Clearance',
                'ResponsibilityOfCRZClearance': 'Responsibility for CRZ Clearance',
                'CurrentStatusCRZClearance': 'Current Status of CRZ Clearance',
                'CRZClearanceDocuments': 'CRZ Clearance Documents',
                'ForestClearance': 'Forest Clearance',
                'ResponsibilityOfForestClearance': 'Responsibility for Forest Clearance',
                'CurrentStatusOfForestClearance': 'Current Status of Forest Clearance',
                'ForestClearanceDocuments': 'Forest Clearance Documents'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Pre_Construction_Stage_Compliance.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response         


# geopackage serializer is not used as no location data is used, so no need to specifically add id
class ConstructionStageComplianceReportView(ListAPIView):
    serializer_class = ConstructionStageComplianceReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, *args, **kwargs):
        try:
            data = ConstructionStage.objects.all().order_by('-id')
            if not data.exists():
                return Response({'message': 'Construction Stage Compliance data not found','status' : '400'},  status=status.HTTP_400_BAD_REQUEST)

            construction_stage_compliance_data = ConstructionStageComplianceReportSerializer(data, many=True).data
                 
            return Response({'message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "construction_stage_compliance_data": construction_stage_compliance_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'message': 'There is no data available for Construction Stage Compliance',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)


class ConstructionStageComplianceExcel(generics.ListAPIView):
    # filter_backends = [DjangoFilterBackend] # discuss if we want filters of package and quarter or not

    def get_queryset(self):
        queryset = ConstructionStage.objects.all()
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id',
            'ConsenttToEstablishOoperate', 'RulesOfConsenttToEstablishOoperate', 'ResponsibilityOfConsenttToEstablishOoperate', 'CurrentStatusOfConsenttToEstablishOoperate', 'ConsenttToEstablishOoperateDocuments',
            'PermissionForSandMiningFromRiverbed', 'RulesOfSandMiningFromRiverbed', 'ResponsibilityOfSandMiningFromRiverbed', 'CurrentStatusOfSandMiningFromRiverbed', 'PermissionForSandMiningFromRiverbedDocuments',
            'PermissionForGroundWaterWithdrawal', 'RulesForGroundWaterWithdrawal', 'ResponsibilityForGroundWaterWithdrawal', 'CurrentStatusOfGroundWaterWithdrawal', 'PermissionForGroundWaterWithdrawalDocuments',
            'AuthorizationForCollectionDisposalManagement', 'RulesForCollectionDisposalManagement', 'ResponsibilityForCollectionDisposalManagement', 'CurrentStatusOfCollectionDisposalManagement', 'AuthorizationForCollectionDisposalManagementDocuments',
            'AuthorizationForSolidWaste', 'RulesForSolidWaste', 'ResponsibilityOfSolidWaste', 'CurrentStatusOfSolidWaste', 'AuthorizationForSolidWasteDocuments',
            'DisposalOfBituminousAndOtherWaste', 'RulesForDisposalOfBituminousAndOtherWaste', 'ResponsibilityOfDisposalOfBituminousAndOtherWaste', 'CurrentStatusOfDisposalOfBituminousAndOtherWaste', 'DisposalOfBituminousAndOtherWasteDocuments',
            'ConsentToDisposalOfsewagefromLabourCamps', 'RulesForDisposalOfsewagefromLabourCamps', 'ResponsibilityOfDisposalOfsewagefromLabourCamps', 'CurrentStatusOfDisposalOfsewagefromLabourCamps', 'ConsentToDisposalOfsewagefromLabourCampsDocuments',
            'PollutionUnderControlCertificate', 'RulesForPollutionUnderControl', 'ResponsibilityOfPollutionUnderControl', 'CurrentStatusPollutionUnderControl', 'PollutionUnderControlCertificateDocuments',
            'RoofTopRainWaterHarvesting', 'RulesForRoofTopRainWaterHarvesting', 'ResponsibilityOfRoofTopRainWaterHarvesting', 'CurrentStatusRoofTopRainWaterHarvesting', 'RoofTopRainWaterHarvestingDocuments'
        )


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'ConsenttToEstablishOoperate': 'Consent to Establish/Operate',
                'RulesOfConsenttToEstablishOoperate': 'Rules for Consent to Establish/Operate',
                'ResponsibilityOfConsenttToEstablishOoperate': 'Responsibility for Consent to Establish/Operate',
                'CurrentStatusOfConsenttToEstablishOoperate': 'Current Status of Consent to Establish/Operate',
                'ConsenttToEstablishOoperateDocuments': 'Consent to Establish/Operate Documents',
                'PermissionForSandMiningFromRiverbed': 'Permission for Sand Mining from Riverbed',
                'RulesOfSandMiningFromRiverbed': 'Rules for Sand Mining from Riverbed',
                'ResponsibilityOfSandMiningFromRiverbed': 'Responsibility for Sand Mining from Riverbed',
                'CurrentStatusOfSandMiningFromRiverbed': 'Current Status of Sand Mining from Riverbed',
                'PermissionForSandMiningFromRiverbedDocuments': 'Sand Mining from Riverbed Documents',
                'PermissionForGroundWaterWithdrawal': 'Permission for Ground Water Withdrawal',
                'RulesForGroundWaterWithdrawal': 'Rules for Ground Water Withdrawal',
                'ResponsibilityForGroundWaterWithdrawal': 'Responsibility for Ground Water Withdrawal',
                'CurrentStatusOfGroundWaterWithdrawal': 'Current Status of Ground Water Withdrawal',
                'PermissionForGroundWaterWithdrawalDocuments': 'Ground Water Withdrawal Documents',
                'AuthorizationForCollectionDisposalManagement': 'Authorization for Collection/Disposal Management',
                'RulesForCollectionDisposalManagement': 'Rules for Collection/Disposal Management',
                'ResponsibilityForCollectionDisposalManagement': 'Responsibility for Collection/Disposal Management',
                'CurrentStatusOfCollectionDisposalManagement': 'Current Status of Collection/Disposal Management',
                'AuthorizationForCollectionDisposalManagementDocuments': 'Collection/Disposal Management Documents',
                'AuthorizationForSolidWaste': 'Authorization for Solid Waste',
                'RulesForSolidWaste': 'Rules for Solid Waste',
                'ResponsibilityOfSolidWaste': 'Responsibility for Solid Waste',
                'CurrentStatusOfSolidWaste': 'Current Status of Solid Waste',
                'AuthorizationForSolidWasteDocuments': 'Solid Waste Documents',
                'DisposalOfBituminousAndOtherWaste': 'Disposal of Bituminous and Other Waste',
                'RulesForDisposalOfBituminousAndOtherWaste': 'Rules for Disposal of Bituminous and Other Waste',
                'ResponsibilityOfDisposalOfBituminousAndOtherWaste': 'Responsibility for Disposal of Bituminous and Other Waste',
                'CurrentStatusOfDisposalOfBituminousAndOtherWaste': 'Current Status of Bituminous/Other Waste Disposal',
                'DisposalOfBituminousAndOtherWasteDocuments': 'Bituminous/Other Waste Disposal Documents',
                'ConsentToDisposalOfsewagefromLabourCamps': 'Consent to Disposal of Sewage from Labour Camps',
                'RulesForDisposalOfsewagefromLabourCamps': 'Rules for Sewage Disposal from Labour Camps',
                'ResponsibilityOfDisposalOfsewagefromLabourCamps': 'Responsibility for Sewage Disposal from Labour Camps',
                'CurrentStatusOfDisposalOfsewagefromLabourCamps': 'Current Status of Sewage Disposal from Labour Camps',
                'ConsentToDisposalOfsewagefromLabourCampsDocuments': 'Sewage Disposal from Labour Camps Documents',
                'PollutionUnderControlCertificate': 'Pollution Under Control Certificate',
                'RulesForPollutionUnderControl': 'Rules for Pollution Under Control',
                'ResponsibilityOfPollutionUnderControl': 'Responsibility for Pollution Under Control',
                'CurrentStatusPollutionUnderControl': 'Current Status of Pollution Under Control',
                'PollutionUnderControlCertificateDocuments': 'Pollution Under Control Certificate Documents',
                'RoofTopRainWaterHarvesting': 'Roof Top Rain Water Harvesting',
                'RulesForRoofTopRainWaterHarvesting': 'Rules for Roof Top Rain Water Harvesting',
                'ResponsibilityOfRoofTopRainWaterHarvesting': 'Responsibility for Roof Top Rain Water Harvesting',
                'CurrentStatusRoofTopRainWaterHarvesting': 'Current Status of Roof Top Rain Water Harvesting',
                'RoofTopRainWaterHarvestingDocuments': 'Roof Top Rain Water Harvesting Documents'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Construction_Stage_Compliance.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response         

            
#    ----- Occupation Health & Safety -----
    
class OccupationalHealthQuarterView(generics.GenericAPIView):
    serializer_class = OccupationalHealthQuarterSeialzier
    def get(self , request , quarter , year):
        data = occupationalHealthSafety.objects.filter(
            quarter=quarter, dateOfMonitoring__year=year).order_by('-id')
        if not data.exists():
            return Response({'Message': 'No data found',
                                },  status=status.HTTP_400_BAD_REQUEST)
        
        occupational_health_data = OccupationalHealthQuarterSeialzier(data, many=True).data
        for feature in occupational_health_data['features']:
                feature['properties']['id'] = feature['id']
                  
        return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                             "training_data": occupational_health_data},
                            status=200)


class OccupationalWellnessPackageExcelDownload(generics.ListAPIView):
    serializer_class = ExcelOccupationalHealthQuarterSeialzier
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = occupationalHealthSafety.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values(
            'id', 'labourCampId', 'labourCampName', 'joiningMedicalCheckup', 'ppeKit', 'trainingToWorkers', 'houseKeeping',
            'powerSupplySystem', 'assemblyArea', 'ambulanceArrangement', 'toiletFacility', 'safeMomentPassage',
            'materialKeepingPractice', 'accidentalCheck', 'safetyGearStatus', 'barricading', 'typeOfIncident',
            'incidentReportingStatus', 'incidentDetails', 'identifiedCauseOfIncident', 'outcome', 'compensationPaid',
            'natureOfAccident', 'manDaysLostCount', 'manDaysLostReason', 'documents', 'photographs', 'remarks'
        )


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'labourCampId': 'Labour Camp ID',
                'labourCampName': 'Labour Camp Name',
                'joiningMedicalCheckup': 'Joining Medical Checkup',
                'ppeKit': 'PPE Kit',
                'trainingToWorkers': 'Training to Workers',
                'houseKeeping': 'House Keeping',
                'powerSupplySystem': 'Power Supply System',
                'assemblyArea': 'Assembly Area',
                'ambulanceArrangement': 'Ambulance Arrangement',
                'toiletFacility': 'Toilet Facility',
                'safeMomentPassage': 'Safe Movement Passage',
                'materialKeepingPractice': 'Material Keeping Practice',
                'accidentalCheck': 'Accidental Check',
                'safetyGearStatus': 'Safety Gear Status',
                'barricading': 'Barricading',
                'typeOfIncident': 'Type of Incident',
                'incidentReportingStatus': 'Incident Reporting Status',
                'incidentDetails': 'Incident Details',
                'identifiedCauseOfIncident': 'Identified Cause of Incident',
                'outcome': 'Outcome',
                'compensationPaid': 'Compensation Paid',
                'natureOfAccident': 'Nature of Accident',
                'manDaysLostCount': 'Man Days Lost Count',
                'manDaysLostReason': 'Reason for Man Days Lost',
                'documents': 'Documents',
                'photographs': 'Photographs',
                'remarks': 'Remarks'
            })

            
            # Get the package filter from the request
            package = request.GET.get('packages', 'all_packages')

            # Create a response with the appropriate content type
            filename = f'Occupational_Wellness_{package}.xlsx'
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class OccupationalHealthPackageView(generics.GenericAPIView):
    serializer_class = OccupationalHealthQuarterSeialzier
    def get(self , request , packages):
        data = occupationalHealthSafety.objects.filter(
            packages=packages ).order_by('-id')
        if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)
        
        occupational_health_data = OccupationalHealthQuarterSeialzier(data, many=True).data
        for feature in occupational_health_data['features']:
                feature['properties']['id'] = feature['id']   
        return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                             "training_data": occupational_health_data},
                            status=200)




class OccupationalWellnessQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')

    class Meta:
        model = occupationalHealthSafety
        fields = ['quarter', 'year']



class OccupationalWellnessQuarterExcelDownload(generics.ListAPIView):
    serializer_class = ExcelOccupationalHealthQuarterSeialzier
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = OccupationalWellnessQuarterFilter

    def get_queryset(self):
        queryset = occupationalHealthSafety.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values(
            'id', 'labourCampId', 'labourCampName', 'joiningMedicalCheckup', 'ppeKit', 'trainingToWorkers', 'houseKeeping',
            'powerSupplySystem', 'assemblyArea', 'ambulanceArrangement', 'toiletFacility', 'safeMomentPassage',
            'materialKeepingPractice', 'accidentalCheck', 'safetyGearStatus', 'barricading', 'typeOfIncident',
            'incidentReportingStatus', 'incidentDetails', 'identifiedCauseOfIncident', 'outcome', 'compensationPaid',
            'natureOfAccident', 'manDaysLostCount', 'manDaysLostReason', 'documents', 'photographs', 'remarks'
        )
        
        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Get the quarter and year filter from the request
            quarter = request.GET.get('quarter', 'all_quarters')
            year = request.GET.get('year', 'all_years')

            # Rename the columns to more appropriate names
            df = df.rename(columns={
                'id': 'ID',
                'labourCampId': 'Labour Camp ID',
                'labourCampName': 'Labour Camp Name',
                'joiningMedicalCheckup': 'Joining Medical Checkup',
                'ppeKit': 'PPE Kit',
                'trainingToWorkers': 'Training to Workers',
                'houseKeeping': 'House Keeping',
                'powerSupplySystem': 'Power Supply System',
                'assemblyArea': 'Assembly Area',
                'ambulanceArrangement': 'Ambulance Arrangement',
                'toiletFacility': 'Toilet Facility',
                'safeMomentPassage': 'Safe Movement Passage',
                'materialKeepingPractice': 'Material Keeping Practice',
                'accidentalCheck': 'Accidental Check',
                'safetyGearStatus': 'Safety Gear Status',
                'barricading': 'Barricading',
                'typeOfIncident': 'Type of Incident',
                'incidentReportingStatus': 'Incident Reporting Status',
                'incidentDetails': 'Incident Details',
                'identifiedCauseOfIncident': 'Identified Cause of Incident',
                'outcome': 'Outcome',
                'compensationPaid': 'Compensation Paid',
                'natureOfAccident': 'Nature of Accident',
                'manDaysLostCount': 'Man Days Lost Count',
                'manDaysLostReason': 'Reason for Man Days Lost',
                'documents': 'Documents',
                'photographs': 'Photographs',
                'remarks': 'Remarks'
            })


            # Create a response with the appropriate content type
            filename = f'Occupational_Wellness_{year}_{quarter}.xlsx'
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


# Training

class TrainnigReportQuarterView(APIView):
    def get(self , request , quarter , year):
   
        data = traning.objects.filter(quarter=quarter, dateOfMonitoring__year=year).order_by('-id')
        if not data.exists():
            return Response({'Message': 'No data found',
                                },  status=status.HTTP_400_BAD_REQUEST)
        
        training_data = TrainnigReportSerializer(data, many=True).data
        for feature in training_data['features']:
                feature['properties']['id'] = feature['id']
                 
        return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                             "training_data": training_data},
                            status=200)


class TrainingQuarterFilter(django_filters.FilterSet):
    year = django_filters.NumberFilter(field_name='dateOfMonitoring__year', label='Year')
    quarter = django_filters.CharFilter(field_name='quarter', label='quarter')

    class Meta:
        model = traning
        fields = ['quarter', 'year']



class TrainningManagementQuarterExcelDownload(generics.ListAPIView):
    serializer_class = TrainnigReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['packages', 'constructionSiteName']
    filterset_class = TrainingQuarterFilter

    def get_queryset(self):
        queryset = traning.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('quarter' , 'packages' , 'dateOfMonitoring',
                  'category' , 'traningTitle' , 'noOfAttends' , 'noOfTimesTrainingConducted',
                  'male','female' , 'inchargePerson', 'traninigInitiatedBy' , 'photographs' , 'documents', 'remarks')

        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)

        else:
                
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
            # Rename columns to more appropriate names
            df = df.rename(columns={
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfMonitoring': 'Monitoring Date',
                'category': 'Training Category',
                'traningTitle': 'Training Title',
                'noOfAttends': 'Number of Attendees',
                'noOfTimesTrainingConducted': 'Number of Sessions',
                'male': 'Number of Male Attendees',
                'female': 'Number of Female Attendees',
                'inchargePerson': 'In-Charge Person',
                'traninigInitiatedBy': 'Training Initiated By',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=TrainningQuater.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response



class TrainnigReportPackageView(APIView):
    def get(self , request , packages):
   
        data = traning.objects.filter(packages=packages ).order_by('-id')
        if not data.exists():
                return Response({'Message': 'No data found',
                                 },  status=status.HTTP_400_BAD_REQUEST)
        
        training_data = TrainnigReportSerializer(data, many=True).data
        for feature in training_data['features']:
                feature['properties']['id'] = feature['id']   
        return Response({'Message': 'data Fetched Successfully',
                            'status' : 'success' , 
                             "training_data": training_data},
                            status=200)


class TrainnigReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = TrainnigReportExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = traning.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('quarter' , 'packages' , 'dateOfMonitoring',
                  'category' , 'traningTitle' , 'noOfAttends' , 'noOfTimesTrainingConducted',
                  'male','female' , 'inchargePerson', 'traninigInitiatedBy' , 'photographs' , 'documents', 'remarks')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)
        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)
            df = df.rename(columns={
                'quarter': 'Quarter',
                'packages': 'Package',
                'dateOfMonitoring': 'Monitoring Date',
                'category': 'Training Category',
                'traningTitle': 'Training Title',
                'noOfAttends': 'Number of Attendees',
                'noOfTimesTrainingConducted': 'Number of Sessions',
                'male': 'Number of Male Attendees',
                'female': 'Number of Female Attendees',
                'inchargePerson': 'In-Charge Person',
                'traninigInitiatedBy': 'Training Initiated By',
                'photographs': 'Photographs',
                'documents': 'Documents',
                'remarks': 'Remarks'
            })


            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=training.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response
        

class MetroLine4View(generics.GenericAPIView):
    serializer_class = MetroLine4AlignmentSerializer

    def get(self, request, *args, **kwargs):
        try:
            MetroLine = MmrdaNew.objects.all()
            serializers = self.get_serializer(MetroLine , many = True).data

            return Response({'status': 'success',
                                'message' : 'data was successfully fetched',
                                'data': serializers},
                                status= 200)
        except :
            return Response({'status' : 'failed',
                            'message' : 'Something went wrong !! Please try again'}, status = 400)
        

class MetroLine4AlignmentReportPackageExcelDownload(generics.ListAPIView):
    serializer_class = MetroLine4AlignmentExcelSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['packages']
    # filterset_class = ConstructionCampFilter

    def get_queryset(self):
        queryset = MmrdaNew.objects.all()
        # Customize the queryset based on your specific requirements
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Use values to convert the queryset to a list of dictionaries
        data = queryset.values('id','quarter','month','dateOfMonitoring','packages','treeID','commanName' ,'botanicalName',
                    'condition', 'noOfTreeCut','actionTaken', 'photographs', 'documents','remarks')


        if not data:
            return JsonResponse({'status':'error','message':'Data Not Found'}, status=400)

        else:
            
            # Create a Pandas DataFrame
            df = pd.DataFrame(data)

            # Create a response with the appropriate content type
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=ExistingTreeManagmentReport.xlsx'

            # Write the DataFrame to the Excel response
            df.to_excel(response, index=False, sheet_name='Sheet1')

            return response


class Package54AlignmentView(generics.GenericAPIView):
    serializer_class = Package54AlignmentSerializer

    def get(self, request, *args, **kwargs):
        try:
            package54 = Package54Alignment.objects.all()
            serializers = Package54AlignmentSerializer(package54 , many = True).data
            return Response({'status': 'success',
                                'message' : 'data was successfully fetched',
                                'data': serializers},
                                status= 200)
        except :
            return Response({'status' : 'failed',
                            'message' : 'Something went wrong !! Please try again'}, status = 400)
      
class package12AlignmentView(generics.GenericAPIView):
    serializer_class = Package12AlignmentSerializer 

    def get(self, request, *args, **kwargs):
        try:
            package12 = Package12Alignment.objects.all()
            serializers = Package12AlignmentSerializer(package12 , many = True).data
            return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': serializers},
                            status= 200)
        except :
            return Response({'status' : 'failed',
                            'message' : 'Something went wrong !! Please try again'}, status = 400)

class package11AlignmentView(generics.GenericAPIView):
    serializer_class = Package11AlignmentSerializer 

    def get(self, request, *args, **kwargs):
        try:
            package11 = Package11Alignment.objects.all()
            serializer = Package11AlignmentSerializer(package11 , many = True).data
            return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': serializer},
                             status= 200)
        except :
            return Response({'status' : 'failed',
                            'message' : 'Something went wrong !! Please try again'}, status = 400)

class package10AlignmentView(generics.GenericAPIView):
    serializer_class = Package10AlignmentSerializer

    def get(self , request):
        try:
            package10 = Package10Alignment.objects.all()
            serializer = Package10AlignmentSerializer(package10 , many = True).data
    
            return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': serializer})
        except :
            return Response({'status' : 'failed',
                            'message' : 'Something went wrong !! Please try again'})

class package09AlignmentView(generics.GenericAPIView):
    serializer_class = Package09AlignmentSerializer  

    def get(self , request):
        try:
            package09 = Package09Alignment.objects.all()
            serializer = Package09AlignmentSerializer(package09 , many = True).data
            return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': serializer},
                             status= 200)
        except :
            return Response({'status' : 'failed',
                            'message' : 'Something went wrong !! Please try again'}, status = 400)

class package08AlignmentView(generics.GenericAPIView):
    serializer_class = Package08AlignmentSerializer  

    def get(self , request):
        
            package08 = Package08Alignment.objects.all()
            serializer = Package08AlignmentSerializer(package08 , many = True)
            return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': serializer.data},
                             status= 200)
        # except :
        #     return Response({'status' : 'failed',
        #                     'message' : 'Something went wrong !! Please try again'}, status = 400)


class MetroStationView(generics.GenericAPIView):
    serializer_class = MetroStationSerializer   

    def get(self , request):
        try:
            metroStation = Station.objects.all()
            serializer = MetroStationSerializer(metroStation , many = True).data
            return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': serializer},
                             status= 200)
        except :
            return Response({'status' : 'failed',
                            'message' : 'Something went wrong !! Please try again'}, status = 400)

class ProjectAffectedTreesView(generics.GenericAPIView):
    serializer_class = ProjectAffectedTreesSerializer
    
    def get(self , request):
        try:
            AffectedTrees = ProjectAffectedTrees.objects.all()
            serializer = ProjectAffectedTreesSerializer(AffectedTrees , many = True).data
            return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data' : serializer},status= 200)
        except :
            return Response({'status' : 'failed',
                            'message' : 'Something went wrong !! Please try again'}, status = 400)


# geopackage serializer is not used as no location data is used, so no need to specifically add id
class PreConstructionStageComplianceReportView(ListAPIView):
    serializer_class = PreConstructionStageComplianceReportSerializer
    parser_classes = [MultiPartParser]

    def get(self, request, packages,*args, **kwargs):
        try:
            data = PreConstructionStage.objects.filter(packages=packages).order_by('-id')
            if not data.exists():
                return Response({'message': 'Pre Construction Stage Compliance data not found','status' : '400'},  status=status.HTTP_400_BAD_REQUEST)

            pre_construction_stage_compliance_data = PreConstructionStageComplianceReportSerializer(data, many=True).data
                 
            return Response({'message': 'data Fetched Successfully',
                            'status' : 'success' , 
                            "pre_construction_stage_compliance_data": pre_construction_stage_compliance_data},
                            status=status.HTTP_200_OK)
        except:
            return Response({'message': 'There is no data available for Pre Construction Stage Compliance',
                            'status' : 'Failed'},
                            status=status.HTTP_400_BAD_REQUEST)


class ExcelWorkbook(generics.GenericAPIView):
    serializer_class = LabourcampReportSerializer

    def get(self, request):
        wb = Workbook()
        ws = wb.active

        ws['A1'] = "Pranav"
        wb.save("sample1")
        

